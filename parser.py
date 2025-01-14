from openpyxl import load_workbook
import csv

data_path = 'data/latest'

wb = load_workbook(f'{data_path}/data.xlsx', read_only=True)

col=[3,3,3,3,4,3,3,5,5]
indice=0

# get the report date
date_list = wb['casi_regioni']['A2'].value.split('/')
report_date = f'{date_list[2]}-{date_list[1]}-{date_list[0]}'

# Write date to file
with open(f'{data_path}/date', 'w') as f:
    f.write(report_date)

# Write data to file
for sheet in wb.worksheets[2:]:
    with open(f'{data_path}/{sheet.title}.csv', 'w', newline="") as f:
        c = csv.writer(f)
        for r in sheet.iter_rows(min_row=1, min_col=1, max_row=1000, max_col=col[indice]):
            c.writerow([cell.value for cell in r])
        indice=indice+1
